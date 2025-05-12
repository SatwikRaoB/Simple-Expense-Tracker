import tkinter as tk
from tkinter import ttk, messagebox
import datetime
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

def load_expenses():
    workbook_name = 'expenses.xlsx'
    try:
        workbook = openpyxl.load_workbook(workbook_name)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Date', 'Description', 'Amount', 'Category'])
    data = list(worksheet.values)
    return pd.DataFrame(data[1:], columns=data[0]) 

def display_expenses(expenses):
    if hasattr(window, 'expense_tree'):
        window.expense_tree.destroy()

    window.expense_tree = ttk.Treeview(window, columns=('Date', 'Description', 'Amount', 'Category'), show='headings')
    window.expense_tree.heading('Date', text='Date')
    window.expense_tree.heading('Description', text='Description')
    window.expense_tree.heading('Amount', text='Amount')
    window.expense_tree.heading('Category', text='Category')
    window.expense_tree.grid(row=6, column=0, columnspan=2, padx=5, pady=5)

    for index, row in expenses.iterrows():
        # Convert amount to string and create a tuple for values
        values = (row['Date'], row['Description'], str(row['Amount']), row['Category'])
        window.expense_tree.insert('', tk.END, values=values)

def add_expense():
    date = date_entry.get()
    description = description_entry.get()
    amount = amount_entry.get()
    category = category_combobox.get()

    workbook_name = 'expenses.xlsx'
    workbook = openpyxl.load_workbook(workbook_name)
    worksheet = workbook.active
    worksheet.append([date, description, float(amount), category]) 
    workbook.save(workbook_name)

    # Update the display and clear input fields
    expenses = load_expenses()
    display_expenses(expenses)
    clear_input_fields()
    load_expenses_amount()

def edit_expense():
    selected = window.expense_tree.focus()
    if not selected:
        messagebox.showerror('Error', 'No expense selected')
        return

    values = window.expense_tree.item(selected, 'values')

    # Prefill input fields
    clear_input_fields()  
    date_entry.insert(0, values[0]) 
    description_entry.insert(0, values[1])
    amount_entry.insert(0, values[2])
    # (Update category_combobox as needed)

    # Modify data in the Excel file (implementation needed)

def clear_input_fields():
    date_entry.delete(0, tk.END)
    description_entry.delete(0, tk.END)
    amount_entry.delete(0, tk.END)
    category_combobox.current(0) 

def show_summary_pie_chart():
    expenses = load_expenses()

    if expenses.empty:
        messagebox.showinfo('Info', 'No expenses to show.')
        return

    # Calculate the total expenses per category
    summary_data = expenses.groupby('Category')['Amount'].sum()

    if summary_data.empty:
        messagebox.showinfo('Info', 'No expenses in selected categories.')
        return

    # Create a pie chart
    plt.figure(figsize=(5, 5))
    plt.pie(summary_data, labels=summary_data.index, autopct='%1.1f%%', startangle=140)
    plt.title('Expense Summary')
    plt.show()

def delete_expense():
    selected = window.expense_tree.focus()
    if not selected:
        messagebox.showerror('Error', 'No expense selected')
        return

    result = messagebox.askyesno('Confirmation', 'Are you sure you want to delete this expense?')
    if result:
        values = window.expense_tree.item(selected, 'values')
        date, description, amount, category = values

        workbook_name = 'expenses.xlsx'
        try:
            # Load expenses into a DataFrame
            expenses_df = load_expenses()

            # Find the row index where the expense exists
            row_index = expenses_df.index[(expenses_df['Date'] == date) &
                                          (expenses_df['Description'] == description) &
                                          (expenses_df['Amount'] == float(amount)) &
                                          (expenses_df['Category'] == category)].tolist()

            if row_index:
                # Drop the expense from the DataFrame
                expenses_df = expenses_df.drop(index=row_index)

                # Save the updated DataFrame to Excel
                expenses_df.to_excel(workbook_name, index=False)

                # Update the display and clear input fields
                display_expenses(expenses_df)
                clear_input_fields()
                load_expenses_amount()
            else:
                messagebox.showerror('Error', 'Expense not found in the worksheet')

        except Exception as e:
            messagebox.showerror('Error', f'Error deleting expense: {str(e)}')

# Create the main window
window = tk.Tk()
window.title("Expense Tracker")

# Input fields
tk.Label(window, text="Date (YYYY-MM-DD):").grid(row=1, column=0, padx=5, pady=5)
date_entry = tk.Entry(window)
date_entry.insert(0, datetime.date.today().strftime('%Y-%m-%d'))  # Prefill today's date
date_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(window, text="Description:").grid(row=2, column=0, padx=5, pady=5)
description_entry = tk.Entry(window)
description_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(window, text="Amount:").grid(row=3, column=0, padx=5, pady=5)
amount_entry = tk.Entry(window)
amount_entry.grid(row=3, column=1, padx=5, pady=5)

tk.Label(window, text="Category:").grid(row=4, column=0, padx=5, pady=5)
categories = ['Food', 'Rent', 'Entertainment', 'Other']
category_combobox = ttk.Combobox(window, values=categories)
category_combobox.current(0)  # Set default selection
category_combobox.grid(row=4, column=1, padx=5, pady=5)

def load_expenses_amount():
    expenses = load_expenses()
    total_expenses_label = tk.Label(window, text=f"Total Expenses: ${expenses['Amount'].sum():,.2f}", font=("Helvetica", 12))
    total_expenses_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

load_expenses_amount()

# Add and Edit buttons
add_button = tk.Button(window, text="Add Expense", command=add_expense)
add_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

delete_button = tk.Button(window, text="Delete Expense", command=delete_expense)
delete_button.grid(row=9, column=0, columnspan=2, padx=5, pady=5)

edit_button = tk.Button(window, text="Edit Expense", command=edit_expense)
edit_button.grid(row=7, column=0, columnspan=2, padx=5, pady=5) 

summary_button = tk.Button(window, text="Show Summary", command=show_summary_pie_chart)
summary_button.grid(row=8, column=0, columnspan=2, padx=5, pady=5)



# Load expenses and display 
expenses = load_expenses()  
display_expenses(expenses)

window.mainloop()
